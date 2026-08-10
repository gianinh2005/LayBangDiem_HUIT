from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
import pandas as pd

from config import EXCEL_FILE


def export_excel(df, summary=None):

    # =============================
    # Chuẩn hóa Học kỳ
    # =============================

    if "Học kỳ" in df.columns:

        df["Học kỳ"] = (
            df["Học kỳ"]
            .astype(str)
            .str.extract(
                r"(HK\d+\s*\(\d{4}\s*-\s*\d{4}\))"
            )[0]
            .fillna("")
        )


    # =============================
    # Đổi cột Đạt thành X
    # =============================

    if "Đạt" in df.columns:

        df["Đạt"] = df["Đạt"].apply(
            lambda x:
                "X"
                if str(x).strip() not in [
                    "",
                    "nan",
                    "None",
                    "0"
                ]
                else ""
        )


    # =============================
    # Xuất nhiều sheet
    # =============================

    with pd.ExcelWriter(
        EXCEL_FILE,
        engine="openpyxl"
    ) as writer:


        df.to_excel(
            writer,
            sheet_name="Bảng điểm",
            index=False
        )


        # -------------------------
        # Sheet Tổng kết
        # -------------------------

        if summary:

            rows = []

            for tr in summary.find_all("tr"):

                cells = tr.find_all(
                    ["td", "th"]
                )

                values = [
                    c.get_text(
                        " ",
                        strip=True
                    )
                    for c in cells
                ]

                if values:
                    rows.append(values)


            pd.DataFrame(rows).to_excel(
                writer,
                sheet_name="Tổng kết",
                index=False,
                header=False
            )


    # =============================
    # Format Excel
    # =============================

    wb = load_workbook(EXCEL_FILE)


    # =============================
    # Sheet Bảng điểm
    # =============================

    ws = wb["Bảng điểm"]


    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )


    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


    # Header

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # Body

    for row in ws.iter_rows(min_row=2):

        for cell in row:

            cell.border = thin_border

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )


    # Tên môn căn trái

    if "Tên môn" in df.columns:

        idx = (
            list(df.columns)
            .index("Tên môn")
            + 1
        )

        for row in range(
            2,
            ws.max_row + 1
        ):

            ws.cell(
                row=row,
                column=idx
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )


    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions


    # Auto width

    for column_cells in ws.columns:

        max_length = 0

        letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value)
            )


        ws.column_dimensions[letter].width = min(
            max(max_length + 3, 10),
            45
        )


    ws.row_dimensions[1].height = 35


    # Width cố định

    preferred_width = {

        "Học kỳ": 18,
        "STT": 8,
        "Mã môn": 15,
        "Tên môn": 45,
        "Lớp": 18,
        "Số TC": 8,
        "Ghi chú": 25

    }


    for name, width in preferred_width.items():

        if name in df.columns:

            idx = (
                list(df.columns)
                .index(name)
                + 1
            )

            ws.column_dimensions[
                get_column_letter(idx)
            ].width = width



    # =============================
    # Format sheet Tổng kết
    # =============================

    if "Tổng kết" in wb.sheetnames:

        ws2 = wb["Tổng kết"]

        for row in ws2.iter_rows():

            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )


        for col in ws2.columns:

            length = max(
                len(str(cell.value))
                if cell.value else 0
                for cell in col
            )

            ws2.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(
                length + 5,
                40
            )


    wb.save(EXCEL_FILE)


    print(
        f"Đã xuất Excel: {EXCEL_FILE}"
    )